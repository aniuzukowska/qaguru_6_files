import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_do_archive():
    file_zip = zipfile.ZipFile('recources/files.zip', 'w')
    file_zip.write('recources/file1.xlsx')
    file_zip.write('recources/file2.csv')
    file_zip.write('recources/file3.pdf')

    lst = file_zip.namelist()
    assert ('recources/file1.xlsx' in lst) \
           and ('recources/file2.csv' in lst) \
           and ('recources/file3.pdf' in lst)
    file_zip.close()


def test_read_and_assert_pdf():
    pdf_reader = PdfReader('recources/file3.pdf')
    text = pdf_reader.pages[0].extract_text()
    assert ('ПРОИЗВОДСТВЕННЫЙ' in text) and ('КАЛЕНДАРЬ' in text) and ('2022' in text)


def test_read_and_assert_csv():
    with open('recources/file2.csv') as csvfile:
        csvfile = csv.reader(csvfile)
        for line_no, line in enumerate(csvfile, 0):
            if line_no == 1:
                assert '14.09.2022' in line[0]


def test_read_and_assert_xlsx():
    workbook = load_workbook('recources/file1.xlsx')
    sheet = workbook.active
    assert 'среда' == sheet.cell(row=2, column=2).value



